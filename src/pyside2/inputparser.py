import csv, os
from report import Report
from openpyxl import Workbook, load_workbook

'''
   InputParser: Functions to parse input
   Properties:
   reports: a list of previously parsed report Objects
   pdcs: a dict of previously parsed PDC map data, with filename keys
'''
class InputParser:

    def __init__(self):
        self.reports = []
        self.pdcs = {}


    #Reads the PDC fuse map CSV and returns a list of the dictionaries
    #Each element in the list is a row of the PDC file, in dict format
    #{CONNECTOR: (component,pin), FUSE:fuse rating}
    def readPDC(self, fileName):

        if fileName:
            contents_list = []
            with open(fileName, mode='r') as csv_file:
                pdcDict = csv.DictReader(csv_file, delimiter=',')
                print("Successfully opened pdc fuse map..")
                name = os.path.basename(fileName)
                for line in pdcDict:
                    contents = {}
                    contents["CONNECTOR"] = (line["CONNECTOR"], line["PIN"])
                    contents["FUSE"] = line["FUSE RATING"]
                    contents_list.append(contents)
                if name not in self.pdcs.keys():
                    self.pdcs[name] = contents_list
                return contents_list

        else:
            print("invalid filename passed to readPDC...")

    #Creates a report object from filename, (from_conn, from_pin), (to_conn, to_pin), csa, description
    #reads the report, stores the object, and returns the data
    def readReport(self, filename, from_labels, to_labels, csa, desc):
        report = Report(filename, from_labels, to_labels, csa, desc)
        self.reports.append(report)
        return report

    #get the current list of report objects
    def getReports(self):
        return self.reports


    #gets first line of worksheet 'filename'
    #and returns it as a list
    def readColumnNames(self, filename):
        if filename:
            wb = load_workbook(filename)
            sheet = wb.active
            names = []
            for col in sheet.iter_cols(1, sheet.max_column, 1, 1, True):
                names.append(col[0])
            return names
